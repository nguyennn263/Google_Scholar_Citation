import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment

def format_excel(_excel_file):

    # Load the Excel file
    workbook = openpyxl.load_workbook(_excel_file)

    # Select the active sheet
    sheet = workbook.active

    # Auto adjust column widths
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                lines = str(cell.value).count('\n') + 1
                if cell_length > max_length:
                    max_length = cell_length / lines
        adjusted_width = (max_length + 2) * 1.2 # Adding padding and scaling factor
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Auto adjust row heights
    alignment = Alignment(horizontal='left', vertical='top')
    for row in sheet.iter_rows():
        max_height = 0
        for cell in row:
            cell.alignment = alignment
            if cell.value:
                lines = str(cell.value).count('\n') + 1  # Counting the number of lines
                cell_height = lines * 14  # Assuming each line has a height of 14 (adjust as needed)
                if cell_height > max_height:
                    max_height = cell_height
        sheet.row_dimensions[row[0].row].height = max_height

    # Save the modified Excel file
    workbook.save(_excel_file)

if __name__ == "__main__":
    excel_file = "names.xlsx"
    format_excel(excel_file)